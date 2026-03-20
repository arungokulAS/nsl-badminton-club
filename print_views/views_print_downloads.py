from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from groups.models import Group
from matches.models import Match
from schedule.models import Court, Round
from teams.models import Team


def download_team_list_xlsx(request):
    teams = Team.objects.all().order_by('team_name')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Teams'
    sheet.append(['#', 'Team', 'Player 1', 'Player 2'])
    for idx, team in enumerate(teams, start=1):
        sheet.append([idx, team.team_name, team.player1_name, team.player2_name])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="team_list.xlsx"'
    return response


def download_group_list_xlsx(request):
    groups = Group.objects.prefetch_related('teams').all().order_by('group_name')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Groups'

    pastel_palette = [
        'FADBD8',
        'D6EAF8',
        'D5F5E3',
        'FCF3CF',
        'E8DAEF',
        'FDEBD0',
    ]
    border = Border(
        left=Side(style='thin', color='BFA8FF'),
        right=Side(style='thin', color='BFA8FF'),
        top=Side(style='thin', color='BFA8FF'),
        bottom=Side(style='thin', color='BFA8FF'),
    )
    header_font = Font(bold=True, color='3D1A78')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    column_positions = [1, 4, 7]
    row_cursor = 1
    group_index = 0
    row_block_height = 0

    for group in groups:
        col = column_positions[group_index % 3]
        if group_index % 3 == 0 and group_index > 0:
            row_cursor += row_block_height + 2
            row_block_height = 0

        teams = list(group.teams.all().order_by('team_name'))
        max_row_height = max(2, len(teams) + 1)
        if max_row_height > row_block_height:
            row_block_height = max_row_height

        header_cell = sheet.cell(row=row_cursor, column=col, value=f'Group {group.group_name}')
        header_cell.font = header_font
        header_cell.alignment = center
        header_cell.fill = PatternFill(fill_type='solid', fgColor=pastel_palette[group_index % len(pastel_palette)])
        header_cell.border = border

        sheet.row_dimensions[row_cursor].height = 22
        sheet.column_dimensions[chr(64 + col)].width = 24

        for idx, team in enumerate(teams, start=1):
            cell = sheet.cell(row=row_cursor + idx, column=col, value=team.team_name)
            cell.alignment = left
            cell.fill = PatternFill(fill_type='solid', fgColor=pastel_palette[group_index % len(pastel_palette)])
            cell.border = border

        for empty_row in range(len(teams) + 1, max_row_height):
            cell = sheet.cell(row=row_cursor + empty_row, column=col, value='')
            cell.fill = PatternFill(fill_type='solid', fgColor=pastel_palette[group_index % len(pastel_palette)])
            cell.border = border

        group_index += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="group_list.xlsx"'
    return response


def download_schedule_xlsx(request):
    standard_round_names = [
        'Group Stage',
        'Qualifier',
        'Pre-Quarter',
        'Quarter',
        'Semi Final',
        'Losers Final',
        'Final',
    ]
    rounds = Round.objects.filter(order__in=[1, 2, 3, 4, 5, 6, 7], name__in=standard_round_names).order_by('order')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Schedule'
    for round_obj in rounds:
        matches = Match.objects.select_related('team1', 'team2', 'court', 'group').filter(
            round=round_obj
        ).order_by('court__id', 'id')
        sheet.append([round_obj.name])
        current_court = None
        for match in matches:
            court_name = match.court.name if match.court else 'Unassigned Court'
            if court_name != current_court:
                current_court = court_name
                sheet.append([f'Court: {current_court}'])
                if round_obj.name == 'Group Stage':
                    sheet.append(['#', 'Group', 'Team 1', 'Team 2'])
                else:
                    sheet.append(['#', 'Team 1', 'Team 2'])
            if round_obj.name == 'Group Stage':
                sheet.append([
                    '',
                    match.group.group_name if match.group else '-',
                    match.team1.team_name if match.team1 else '-',
                    match.team2.team_name if match.team2 else '-',
                ])
            else:
                sheet.append([
                    '',
                    match.team1.team_name if match.team1 else '-',
                    match.team2.team_name if match.team2 else '-',
                ])
        sheet.append([])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="schedule.xlsx"'
    return response


def _safe_filename(value):
    cleaned = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in value.strip())
    return cleaned or 'schedule'


def download_schedule_court_xlsx(request, round_id, court_key):
    round_obj = get_object_or_404(Round, id=round_id)
    if court_key == 'unassigned':
        court = None
        court_name = 'Unassigned Court'
        matches = Match.objects.select_related('team1', 'team2', 'court', 'group').filter(
            round=round_obj, court__isnull=True
        ).order_by('id')
    else:
        try:
            court_id = int(court_key)
        except (TypeError, ValueError):
            return HttpResponse(status=404)
        court = get_object_or_404(Court, id=court_id)
        court_name = court.name
        matches = Match.objects.select_related('team1', 'team2', 'court', 'group').filter(
            round=round_obj, court=court
        ).order_by('id')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f'{round_obj.name} - {court_name}'[:31]
    sheet.append([round_obj.name])
    sheet.append([f'Court: {court_name}'])
    if round_obj.name == 'Group Stage':
        sheet.append(['#', 'Group', 'Team 1', 'Team 2'])
        for idx, match in enumerate(matches, start=1):
            sheet.append([
                idx,
                match.group.group_name if match.group else '-',
                match.team1.team_name if match.team1 else '-',
                match.team2.team_name if match.team2 else '-',
            ])
    else:
        sheet.append(['#', 'Team 1', 'Team 2'])
        for idx, match in enumerate(matches, start=1):
            sheet.append([
                idx,
                match.team1.team_name if match.team1 else '-',
                match.team2.team_name if match.team2 else '-',
            ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"schedule_{_safe_filename(round_obj.name)}_{_safe_filename(court_name)}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
